# user input automation and scraping
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException

# for sleep() and date/time file naming
import time
from datetime import datetime

# for creating and modifying xlsx files with data scraped
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# for file directory
import os

# for auto-mailing
import smtplib

while True:
    try:
        term = input("Enter semester (fall, winter, spring, summer): ")
        if term.lower() not in ["fall", "winter", "spring", "summer"]:
            raise ValueError("Invalid semester entered.")
        break
    except ValueError as e:
        print(e)

while True:
    try:
        show_open = input("Show Open classes only (yes/no)?: ")
        if show_open.lower() not in ['yes', 'no']:
            raise ValueError("Please enter 'yes' or 'no'")
        break
    except ValueError as e:
        print(e)

while True:

    url = 'https://globalsearch.cuny.edu/CFGlobalSearchTool/search.jsp'

    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)

    driver = webdriver.Chrome(options=options)
    driver.get(url)

    print("Page #1:")
    time.sleep(3)

    # Page 1
    # Institution
    driver.find_element(By.CSS_SELECTOR, '#CTY01').click()
    print("clicked City College of New York")
    time.sleep(1)
    # Term: choose SPRING[4], SUMMER[3], FALL[2]
    term_options = []
    if term.lower() == 'spring':
        term_options.append(4)
        term_options.append(21)
    elif term.lower() == 'summer':
        term_options.append(3)
        term_options.append(14)
    elif term.lower() == 'fall':
        term_options.append(2)
        term_options.append(19)
    driver.find_element(By.XPATH, '//*[@id="t_pd"]/option[{}]'
                        .format(term_options[0])).click()
    print("clicked", term.upper(), "semester")
    time.sleep(1)
    # Next button
    driver.find_element(By.CLASS_NAME, 'SSSBUTTON_CONFIRMLINK').click()
    print("clicked 'NEXT' button" + '\n')

    print("Page #2:")
    time.sleep(3)

    # Page 2
    # Subject: SPRING[21], SUMMER[14], FALL[19]
    driver.find_element(By.XPATH, '//*[@id="subject_ld"]/option[{}]'
                        .format(term_options[1])).click()
    print("clicked Computer Science subject")
    time.sleep(1)
    # Course Career
    driver.find_element(By.XPATH, '//*[@id="courseCareerId"]/option[4]').click()
    print("clicked Undergraduate")
    time.sleep(1)
    # Shown Open Classes Only
    if show_open.lower() == 'no':
        driver.find_element(By.XPATH, '//*[@id="open_classId"]').click()
        print("show ALL Classes")
        time.sleep(1)
    else:
        print("show OPEN Classes only")
    # Search
    driver.find_element(By.ID, 'btnGetAjax').click()
    print("clicked 'SEARCH' button" + '\n')
    
    print("Page #3:")
    time.sleep(3)

    # Page 3
    # data lists
    cs_title_list = []
    cs_section_list = []
    cs_DnT_list = []
    cs_availability_list = []
    cs_instructors_list = []
    cs_mode_list = []
    # Show courses
    driver.find_element(By.ID, 'imageDivLink_inst0').click()
    print("showing ALL Courses..." + '\n')
    time.sleep(1)

    # Show all classes
    i = 0 # courses
    j = 2 # classes in a course. 2 is starting index of row
    # Course/Class counters
    cs_courses = 1 
    cs_classes = 0
    cs_classes_sum = 0
    # Clicks course drop down
    while True:
        try:
            driver.find_element(By.ID, 'imageDivLink{}'.format(i)).click()
            while True:
                try: 
                    # Find class' title
                    cs_title = driver.find_element(By.XPATH, '//*[@id="contentDivImg_inst0"]/table[{}]/tbody/tr/td/b/span[@class="cunylite_LABEL"]'
                                                .format(cs_courses)).text
                    # Find class' section
                    cs_section = driver.find_element(By.XPATH, '//*[@id="contentDivImg{}"]/table/tbody/tr[{}]/td[3]'
                                                    .format(i,j)).text
                    # Find class' Days & Times
                    cs_DnT = driver.find_element(By.XPATH, '//*[@id="contentDivImg{}"]/table/tbody/tr[{}]/td[4]'
                                                    .format(i,j)).text
                    # Find class' availability
                    cs_availability = driver.find_element(By.XPATH, '//*[@id="contentDivImg{}"]/table/tbody/tr[{}]/td[9]/img'
                                                    .format(i,j))
                    availability = cs_availability.get_attribute('title')
                    # Find class' instructors
                    cs_instructors = driver.find_element(By.XPATH, '//*[@id="contentDivImg{}"]/table/tbody/tr[{}]/td[6]'
                                                    .format(i,j)).text
                    # Find class' instruction mode
                    cs_mode = driver.find_element(By.XPATH, '//*[@id="contentDivImg{}"]/table/tbody/tr[{}]/td[7]'
                                                    .format(i,j)).text
                    # if availability == 'Open':
                    cs_title_list.append(cs_title)
                    cs_section_list.append(cs_section)
                    cs_DnT_list.append(cs_DnT)
                    cs_availability_list.append(availability)
                    cs_instructors_list.append(cs_instructors)
                    cs_mode_list.append(cs_mode)
                    # Next class
                    cs_classes += 1
                    cs_classes_sum += 1
                    j += 1
                except NoSuchElementException:
                    # print(cs_classes, 'classes founded.')
                    break
            print(cs_classes, 'class(es) found in:' + cs_title)
            i += 1  # Next course
            cs_courses += 1 # next course title
            j = 2   # reset starting index of classes row
            cs_classes = 0
        except NoSuchElementException:
            print('\n' + "Finished collecting all CS courses data.")
            print(i, 'CS Courses found.')
            print(cs_classes_sum, 'CS Class Sections in total.' + '\n')
            break

    # Import data to xlsx
    header = ["Course", "Section", "Days & Time", "Availability", "Instructors", "Mode"]
    data = []
    for i in range(len(cs_title_list)):
        data.append((cs_title_list[i], cs_section_list[i], cs_DnT_list[i], cs_availability_list[i], cs_instructors_list[i], cs_mode_list[i]))

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Select the active worksheet
    worksheet = workbook.active

    # Set the header row
    header = ["Course", "Section", "Days & Time", "Availability", "Instructors", "Mode"]
    worksheet.append(header)

    # Add the data
    for i in range(len(cs_title_list)):
        row = [cs_title_list[i], cs_section_list[i], cs_DnT_list[i], cs_availability_list[i], cs_instructors_list[i], cs_mode_list[i]]
        worksheet.append(row)

    # Get current time and date
    now = datetime.now()
    current_date_time = now.strftime("%Y%m%d-%H%M")

    xlsx_f = str(current_date_time) + "_" + term.upper() + "_CCNY_CS_Classes.xlsx"

    # specify directory/file path
    dir = 'C:/Users/Zed/Documents/Code/repos/cunyf_enrollme/class-status-logs'
    fp = os.path.join(dir, xlsx_f)
    # Save the workbook
    workbook.save(fp)

    print("Created XLSX file named:" + '\n\t' + xlsx_f)

    wb = load_workbook(fp)
    ws = wb['Sheet']

    # Update column width
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10

    # Update row height
    # enable text wrapping for the cells
    for cell in ws['C:C']:
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    for cell in ws['E:E']:
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
    # set the row height to auto adjust for all rows
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        row[0].parent.auto_size = True

    # WANTED courses      
    desired_courses = ['  CSC 30100 - Scientific Prgrmng',
                       '  CSC 30400 - Intro to Theoretical Comp Sci',
                       '  CSC 31800 - Internet Programming',
                       '  CSC 33500 - Programming Language Paradigms',
                       '  CSC 33600 - Database Systems',
                       '  CSC 34200 - Computer Organization',
                       '  CSC 34300 - Computer Systems Design Lab',
                       '  CSC 41200 - Computer Networks',
                       '  CSC 44700 - Introduction to Machine Learni',
                       '  CSC 44800 - Artificial Intelligence',
                       '  CSC 45600 - Topics in Software Engineering',
                       '  CSC 47300 - Web Site Design',
                    #    '',
                    #    '',
                    #    '',
                       ]
    # Aqua = classes i want to take
    fillAqua = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    for row in ws.iter_rows():
        if row[0].value in desired_courses:
            for cell in row:
                cell.fill = fillAqua

    # CURRENTLY taking courses
    current_courses = ['  CSC 34200 - Computer Organization',
                       '  CSC 34300 - Computer Systems Design Lab',
                       '  CSC 38000 - Computer Security',
                       '  CSC 32200 - Software Engineering',
                    #    '',
                       ]
    # orange = currently taking
    fillOrange = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
    for row in ws.iter_rows():
        if row[0].value in current_courses:
            for cell in row:
                cell.fill = fillOrange

    # NEEDED courses      
    needed_courses = ['  CSC 30100 - Scientific Prgrmng',
                       '  CSC 30400 - Intro to Theoretical Comp Sci',
                    #    '  CSC 31800 - Internet Programming',
                       '  CSC 33500 - Programming Language Paradigms',
                       '  CSC 33600 - Database Systems',
                    #    '  CSC 34200 - Computer Organization',
                    #    '  CSC 34300 - Computer Systems Design Lab',
                       '  CSC 59866 - Senior Project I',
                       '  CSC 59867 - Senior Project II',
                    #    '',
                       ]
    # purple = classes i need to take
    fillPurple = PatternFill(start_color='800080', end_color='800080', fill_type='solid')
    for row in ws.iter_rows():
        if row[0].value in needed_courses:
            for cell in row:
                cell.fill = fillPurple

    # NEXT semester planned classes
    next_courses = [#'  CSC 30100 - Scientific Prgrmng',
                       '  CSC 30400 - Intro to Theoretical Comp Sci',
                       '  CSC 31800 - Internet Programming',
                       '  CSC 41200 - Computer Networks',
                       '  CSC 33500 - Programming Language Paradigms',
                       '  CSC 33600 - Database Systems',
                    #    '  CSC 34200 - Computer Organization',
                    #    '  CSC 34300 - Computer Systems Design Lab',
                       '  CSC 59866 - Senior Project I',
                    #    '  CSC 59867 - Senior Project II',
                    #    '',
                       ]
    # purple = classes i need to take
    fillYellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for row in ws.iter_rows():
        if row[0].value in next_courses:
            for cell in row:
                cell.fill = fillYellow

    # TAKEN courses
    taken_courses = ['  CSC 10300 - Intrmd Cmptr Progrm',
                     '  CSC 10400 - Discrete Structrs 1',
                     '  CSC 11300 - Programming Lang',
                     '  CSC 21100 - Fundamentals Computer Systems',
                     '  CSC 21200 - Data Structures',
                     '  CSC 21700 - Prob & Stat For Csc',
                     '  CSC 22000 - Algorithms',
                     '  CSC 22100 - Software Design Laboratory',
                     '  CSC 33200 - Operating Systems',
                     ]
    # gray = <courses already taken>
    fillGray = PatternFill(start_color='B2BEB5', end_color='B2BEB5', fill_type='solid')
    # paints whole row Red if its a course already taken
    for row in ws.iter_rows():
        if row[0].value in taken_courses:
            for cell in row:
                cell.fill = fillGray

    # green = "Open", blue = "Closed"
    fillGreen = PatternFill(start_color='70bf22', end_color='70bf22', fill_type='solid')
    fillBlue = PatternFill(start_color='96bfec', end_color='96bfec', fill_type='solid')
    for row in ws.iter_rows():
        for cell in row:
            if 'Open' in cell.value:
                cell.fill = fillGreen
            elif 'Closed' in cell.value:
                cell.fill = fillBlue

    # save and close updated xlsx file
    wb.save(fp)
    wb.close()
    print("XLSX file contents updated!" + "\n")

    # auto email
    # set up the SMTP server
    # smtp_server = 'smtp.gmail.com'
    # smtp_port = 587  # use 465 for SSL/TLS encryption
    # smtp_username = 'vmonet2022@gmail.com'  # replace with your email address
    # smtp_password = 'your_email_password'   # replace with your email password
    # sender_email = 'your_email@gmail.com'   # replace with your email address
    # recipient_email = 'recipient_email@example.com'  # replace with the recipient's email address

    print("Data scraped successfully!")
    driver.quit()
    print("Starting up again in 60 seconds.")
    time.sleep(60)
